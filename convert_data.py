import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
import os
import re

DYNAMIC = 1
POSTHOC = 2

QUICK = 1
REGULAR = 2

demographics_multi = [
    'demographics_gender.csv',
    'demographics_race.csv'
]
demographics_text = [
    'demographics_age.csv'
]
dynamic_filename = re.compile(r'lie_truth_dynamic_([0-9]*)\.csv')
lock_in_filename = re.compile(r'lie_truth_lock_in_([0-9]*)\.csv')
lock_in_decision = re.compile(r'lie_truth_lock_in_decision_([0-9]*)\.csv')
confidence_file_name = re.compile(r'confidence_([0-9]*)\.csv')
dichotomous_legend = [
    'Lie',
    'Truth'
]
video_id_extract = re.compile(r'videos/([0-9]*)/alibi([0-9])_control_trimmed\.webm')

def extract_video_id(matcher, filename):
    video_id = None
    search = matcher.search(filename)

    if search is not None:
        video_id = int(search.groups()[0])

    return video_id

def setup_video_struct(videos: dict, video_id: int) -> dict:
    if 'videos' not in videos:
        videos['videos'] = {}

    if video_id not in videos['videos'].keys():
        videos['videos'][video_id] = {}

    return videos

def write_data_to_cells(worksheet, row, data, starting_column = 1, bold = False):
    column = starting_column
    for d in data:
        worksheet.cell(row, column, d)
        if bold:
            worksheet.cell(row, column).font = Font(bold=True)
        column += 1

def create_video_id(v_id, key, variable):
    additional = ''
    if key == 'lie':
        additional = 'L'
    elif key == 'truth':
        additional = 'T'

    return f'V{v_id}{additional}_{variable}'

def write_and_increment_column(worksheet, row, current_count, v_id, key, variable):
    worksheet.cell(row, current_count, create_video_id(v_id, key, variable))
    return current_count + 1

def write_to_dynamic_xlsx(all_participants):
    workbook = Workbook()
    worksheet = workbook.create_sheet('Data')

    worksheet.cell(1, 1, 'ParticipantID')
    column_id = 2

    max_vid_counts = {}
    column_lookup = {}

    for p_id in sorted(all_participants.keys()):
        for vnum_id in all_participants[p_id]['videos'].keys():
            video_match = video_id_extract.match(all_participants[p_id]['videos'][vnum_id]['video_filename'])

            v_id = int(video_match.group(1))
            lt_condition = int(video_match.group(2))
            lt_str = 'truth' if lt_condition == 1 else 'lie'

            if v_id not in max_vid_counts:
                max_vid_counts[v_id] = {}
                max_vid_counts[v_id]['lie'] = 0
                max_vid_counts[v_id]['truth'] = 0

            if len(all_participants[p_id]['videos'][vnum_id]['dynamic_decisions']) > max_vid_counts[v_id][lt_str]:
                max_vid_counts[v_id][lt_str] = len(all_participants[p_id]['videos'][vnum_id]['dynamic_decisions'])

    for v_id in max_vid_counts.keys():
        for i in range(max_vid_counts[v_id]['lie']):
            column_lookup[create_video_id(v_id, 'lie', f'Dec{i+1}')] = column_id
            column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', f'Dec{i+1}')

        for i in range(max_vid_counts[v_id]['lie']):
            column_lookup[create_video_id(v_id, 'lie', f'Dec{i+1}_RT')] = column_id
            column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', f'Dec{i+1}_RT')

        for i in range(max_vid_counts[v_id]['truth']):
            column_lookup[create_video_id(v_id, 'truth', f'Dec{i+1}')] = column_id
            column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', f'Dec{i+1}')

        for i in range(max_vid_counts[v_id]['truth']):
            column_lookup[create_video_id(v_id, 'truth', f'Dec{i+1}_RT')] = column_id
            column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', f'Dec{i+1}_RT')

    print('Found largest amounts...')

    row = 2

    for p_id in sorted(all_participants.keys()):
        if (p_id+1) % 2 == 1:
            continue

        worksheet.cell(row, 1, p_id)

        for vnum_id in all_participants[p_id]['videos'].keys():
            video_match = video_id_extract.match(all_participants[p_id]['videos'][vnum_id]['video_filename'])

            v_id = int(video_match.group(1))
            lt_condition = int(video_match.group(2))
            lt_str = 'truth' if lt_condition == 1 else 'lie'

            dec_id = 1
            for decision in all_participants[p_id]['videos'][vnum_id]['dynamic_decisions']:
                worksheet.cell(row, column_lookup[create_video_id(v_id, lt_str, f'Dec{dec_id}')], decision[1])
                worksheet.cell(row, column_lookup[create_video_id(v_id, lt_str, f'Dec{dec_id}_RT')], decision[0])
                dec_id += 1

        row += 1

    workbook.remove(workbook['Sheet'])
    workbook.save(f'reports/AllDecisions.xlsx')

def write_to_xlsx(participant, data):
    workbook = Workbook()

    if 'demographics_age' in data and 'demographics_race' in data and 'demographics_gender' in data:
        # Write all demographics to first sheet
        demographics_worksheet = workbook.create_sheet('Demographics')

        write_data_to_cells(demographics_worksheet, 1, ('Age', data['demographics_age']))
        write_data_to_cells(demographics_worksheet, 2, ('Race', f"{data['demographics_race'][1].strip()} ({data['demographics_race'][0]})"))
        write_data_to_cells(demographics_worksheet, 3, ('Gender', f"{data['demographics_gender'][1].strip()} ({data['demographics_gender'][0]})"))


    for v_id in sorted(data['videos'].keys()):
        video_worksheet = workbook.create_sheet(f'Video {v_id+1}')

        write_data_to_cells(video_worksheet, 1, ('Video path', data['videos'][v_id]['video_filename']))

        # Dynamic data
        video_worksheet.cell(4, 1, "Lock In Decisions")
        video_worksheet.cell(4, 1).font = Font(bold=True)
        write_data_to_cells(video_worksheet, 5, ('Timestamp', 'Interim decision', 'Velocity'), bold=True)

        current_row = 6
        for d in data['videos'][v_id]['dynamic_decisions']:
            write_data_to_cells(video_worksheet, current_row, d)
            current_row += 1

        current_row +=1
        write_data_to_cells(video_worksheet, current_row, ('Final timestamp', 'Final decision'), bold=True)

        current_row += 1
        write_data_to_cells(video_worksheet, current_row, data['videos'][v_id]['dynamic_final'])

        # Lock in data (only write if we actually have data)
        if 'dichotomous_final' in data['videos'][v_id]:
            video_worksheet.cell(4, 5, "Final Decisions")
            video_worksheet.cell(4, 5).font = Font(bold=True)
            write_data_to_cells(video_worksheet, 5, ('Timestamp', 'Interim decision'), starting_column=5, bold=True)

            current_row = 6
            for d in data['videos'][v_id]['dichotomous_decisions']:
                write_data_to_cells(video_worksheet, current_row, d, starting_column=5)
                current_row += 1

            current_row +=1
            write_data_to_cells(video_worksheet, current_row, ('Final timestamp', 'Final decision'), starting_column=5, bold=True)

            current_row += 1
            write_data_to_cells(video_worksheet, current_row, data['videos'][v_id]['dichotomous_final'], starting_column=5)

        if 'confidence_final' in data['videos'][v_id]:
            video_worksheet.cell(4, 8, "Confidence Decisions")
            video_worksheet.cell(4, 8).font = Font(bold=True)
            write_data_to_cells(video_worksheet, 5, ('Timestamp', 'Interim decision'), starting_column=8, bold=True)

            current_row = 6
            for d in data['videos'][v_id]['confidence_decisions']:
                write_data_to_cells(video_worksheet, current_row, d, starting_column=8)
                current_row += 1

            current_row += 1
            write_data_to_cells(video_worksheet, current_row, ('Final timestamp', 'Final decision'), starting_column=8, bold=True)

            current_row += 1
            write_data_to_cells(video_worksheet, current_row, data['videos'][v_id]['confidence_final'], starting_column=8)

    # Write the workbook to disk
    if not Path('reports/').exists():
        Path('reports/').mkdir()

    workbook.remove(workbook['Sheet'])
    workbook.save(f'reports/Participant {participant}.xlsx')

if __name__ == '__main__':
    all_participants = {}
    print('Converting data files into XLXS files...')

    subdirs = [d[0] for d in os.walk('output/')]

    for d in subdirs:
        split_path = d.split('/')

        if len(split_path) == 2 and len(split_path[1]) == 0:
            continue

        participant_id = int(split_path[1])
        videos = {}

        print(f'Converting participant data in "{d}"')

        # Extract all of the data from the participant files
        for f in os.listdir(d):
            dynamic_match = dynamic_filename.match(f)
            lock_in_match = lock_in_filename.match(f)
            if dynamic_match or lock_in_match:
                matcher = dynamic_match if dynamic_match else lock_in_match
                video_id = extract_video_id(dynamic_filename if dynamic_match else lock_in_filename, f)

                if video_id is not None:
                    videos = setup_video_struct(videos, video_id)

                    with open(d + '/' + f, 'r') as csv_file:
                        csv_data = csv.DictReader(csv_file)

                        videos['videos'][video_id]['dynamic_decisions'] = [(int(d['timestamp']), int(d['value']), float(d['velocity'])) for d in csv_data if d['type'] == 'decision']
                        csv_file.seek(0)
                        videos['videos'][video_id]['dynamic_final'] = [(int(d['timestamp']), int(d['value'])) for d in csv_data if d['type'] == 'final'][0]
                        csv_file.seek(0)
                        videos['videos'][video_id]['video_filename'] = [d['value'] for d in csv_data if d['type'] == 'path'][0]
            elif lock_in_decision.match(f):
                video_id = extract_video_id(lock_in_decision, f)

                if video_id is not None:
                    videos = setup_video_struct(videos, video_id)

                    with open(d + '/' + f, 'r') as csv_file:
                        csv_data = csv.DictReader(csv_file)

                        videos['videos'][video_id]['dichotomous_decisions'] = [(int(d['timestamp']), int(d['value'])) for d in csv_data if d['type'] == 'decision']
                        csv_file.seek(0)
                        videos['videos'][video_id]['dichotomous_final'] = [(int(d['timestamp']), int(d['value'])) for d in csv_data if d['type'] == 'final'][0]
            elif confidence_file_name.match(f):
                video_id = extract_video_id(confidence_file_name, f)

                if video_id is not None:
                    videos = setup_video_struct(videos, video_id)

                    with open(d + '/' + f, 'r') as csv_file:
                        csv_data = csv.DictReader(csv_file)
                        videos['videos'][video_id]['confidence_decisions'] = [(int(d['timestamp']), int(d['value'])) for d in csv_data if d['type'] == 'decision']
                        csv_file.seek(0)
                        videos['videos'][video_id]['confidence_final'] = [(int(d['timestamp']), int(d['value'])) for d in csv_data if d['type'] == 'final'][0]

            for d_file in demographics_multi:
                if Path(d + '/' + d_file).exists():
                    with open(d + '/' + d_file, 'r') as csv_file:
                        csv_data = csv.DictReader(csv_file)
                        csv_obj = next(csv_data)
                        videos[d_file.split('.')[0]] = (int(csv_obj['index']), csv_obj['label'])

            for d_file in demographics_text:
                if Path(d + '/' + d_file).exists():
                    with open(d + '/' + d_file, 'r') as csv_file:
                        csv_data = csv.DictReader(csv_file)
                        csv_obj = next(csv_data)
                        videos[d_file.split('.')[0]] = csv_obj['text']

        all_participants[participant_id] = videos
        #write_to_xlsx(participant_id, videos)

    write_to_dynamic_xlsx(all_participants)
