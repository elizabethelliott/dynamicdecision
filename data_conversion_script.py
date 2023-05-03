import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
import os
import re

lock_in_time = re.compile(r'lie_truth_lock_in_([0-9]*)\.csv')
dynamic_filename = re.compile(r'lie_truth_dynamic_([0-9]*)\.csv')
final_decision_filename = 'lie_truth_lock_in_decision_{file_id}.csv'
confidence_filename = 'confidence_{file_id}.csv'

# First match is the video id #, the second match is lie/truth (1 = truth, 2 = lie)
video_id_matcher = re.compile(r'videos/([0-9]*)/alibi([0-9])_control_trimmed\.webm')

def extract_video_id_and_condition(matcher, video_path):
    video_id = None
    lie_or_truth = 0
    search = matcher.search(video_path)

    if search is not None:
        video_id = int(search.groups()[0])
        lie_or_truth = int(search.groups()[1])

    return video_id, lie_or_truth

def setup_video_struct(videos: dict, video_id: int, lie_or_truth: int) -> dict:
    if 'videos' not in videos:
        videos['videos'] = {}

    if video_id not in videos['videos'].keys():
        videos['videos'][video_id] = {}

    if lie_or_truth == 2 and 'lie' not in videos['videos'][video_id]: # Lie
        videos['videos'][video_id]['lie'] = []
    elif lie_or_truth == 1 and 'truth' not in videos['videos'][video_id]: # Truth
        videos['videos'][video_id]['truth'] = []

    return videos

def write_data_to_cells(worksheet, row, data, starting_column = 1, bold = False):
    column = starting_column
    for d in data:
        worksheet.cell(row, column, d)
        if bold:
            worksheet.cell(row, column).font = Font(bold=True)
        column += 1

def create_video_id(v_id, key, variable):
    additional = ''
    if key == 'lie':
        additional = 'L'
    elif key == 'truth':
        additional = 'T'

    return f'V{v_id}{additional}_{variable}'

def write_and_increment_column(worksheet, row, current_count, v_id, key, variable):
    worksheet.cell(row, current_count, create_video_id(v_id, key, variable))
    return current_count + 1

def write_to_single_xlsx(data):
    workbook = Workbook()
    worksheet = workbook.create_sheet('Data')

    worksheet.cell(1, 1, 'ParticipantID')

    # Create video look up table
    sorted_video_ids = sorted(data['videos'].keys())
    column_id = 2

    video_id_lookup = {}
    for v_id in sorted_video_ids:
        video_id_lookup[create_video_id(v_id, 'lie', 'LockIn')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', 'LockIn')
        video_id_lookup[create_video_id(v_id, 'lie', 'LockInRT')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', 'LockInRT')
        video_id_lookup[create_video_id(v_id, 'lie', 'Final')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', 'Final')
        video_id_lookup[create_video_id(v_id, 'lie', 'FinalRT')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', 'FinalRT')
        video_id_lookup[create_video_id(v_id, 'lie', 'Confidence')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', 'Confidence')
        video_id_lookup[create_video_id(v_id, 'lie', 'ConfidenceRT')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'lie', 'ConfidenceRT')

        video_id_lookup[create_video_id(v_id, 'truth', 'LockIn')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', 'LockIn')
        video_id_lookup[create_video_id(v_id, 'truth', 'LockInRT')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', 'LockInRT')
        video_id_lookup[create_video_id(v_id, 'truth', 'Final')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', 'Final')
        video_id_lookup[create_video_id(v_id, 'truth', 'FinalRT')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', 'FinalRT')
        video_id_lookup[create_video_id(v_id, 'truth', 'Confidence')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', 'Confidence')
        video_id_lookup[create_video_id(v_id, 'truth', 'ConfidenceRT')] = column_id
        column_id = write_and_increment_column(worksheet, 1, column_id, v_id, 'truth', 'ConfidenceRT')

    participant_lookup = {}
    row = 2
    for v_id in sorted_video_ids:
        for cond in sorted(data['videos'][v_id].keys()):
            for d in data['videos'][v_id][cond]:
                if d[0] in participant_lookup.keys():
                    selected_row = participant_lookup[d[0]]
                else:
                    selected_row = row
                    participant_lookup[d[0]] = row
                    row += 1

                worksheet.cell(selected_row, 1, d[0])
                worksheet.cell(selected_row, video_id_lookup[create_video_id(v_id, cond, 'LockIn')], d[1])
                worksheet.cell(selected_row, video_id_lookup[create_video_id(v_id, cond, 'LockInRT')], d[2])
                worksheet.cell(selected_row, video_id_lookup[create_video_id(v_id, cond, 'Final')], d[3])
                worksheet.cell(selected_row, video_id_lookup[create_video_id(v_id, cond, 'FinalRT')], d[4])
                worksheet.cell(selected_row, video_id_lookup[create_video_id(v_id, cond, 'Confidence')], d[5])
                worksheet.cell(selected_row, video_id_lookup[create_video_id(v_id, cond, 'ConfidenceRT')], d[6])

    workbook.remove(workbook['Sheet'])
    workbook.save(f'reports/DataSingle.xlsx')

def write_to_xlsx(data):
    workbook = Workbook()

    for v_id in sorted(data['videos'].keys()):
        for cond in sorted(data['videos'][v_id].keys()):
            video_worksheet = workbook.create_sheet(f'Video {v_id}{"L" if cond == "lie" else "T"}')

            row = 1
            write_data_to_cells(video_worksheet, row, ('ParticipantID', 'Lock In #', 'Lock In RT', 'Final #', 'Final RT', 'Confidence', 'Confidence RT'), bold=True)
            for d in data['videos'][v_id][cond]:
                row += 1
                write_data_to_cells(video_worksheet, row, d)

    # Write the workbook to disk
    if not Path('reports/').exists():
        Path('reports/').mkdir()

    workbook.remove(workbook['Sheet'])
    workbook.save(f'reports/Data.xlsx')

if __name__ == '__main__':
    print('Converting data files into XLXS files...')

    subdirs = [d[0] for d in os.walk('output/')]

    videos = {}

    for d in subdirs:
        split_path = d.split('/')

        if len(split_path) == 2 and len(split_path[1]) == 0:
            continue

        participant_id = int(split_path[1])

        # if participant_id % 2 != 1:
        #     continue
        
        print(f'Reading participant data in "{d}"')

        # Extract all of the data from the participant files
        for f in os.listdir(d):
            dynamic_match = dynamic_filename.match(f)
            lock_in_match = lock_in_time.match(f)
            if dynamic_match or lock_in_match:
                matcher = dynamic_match if dynamic_match else lock_in_match
                file_id = matcher.group(1)
                with open(d + '/' + f, 'r') as csv_file:
                    csv_data = csv.DictReader(csv_file)

                    path = [d['value'] for d in csv_data if d['type'] == 'path'][0]
                    csv_file.seek(0)

                    video_id, lie_or_truth = extract_video_id_and_condition(video_id_matcher, path)

                    if video_id is not None:
                        videos = setup_video_struct(videos, video_id, lie_or_truth)

                        lt_str = 'truth' if lie_or_truth == 1 else 'lie'
                        data_struct = None

                        # Get all of the dynamic decisions
                        dynamic_decisions = [(d['value'], d['timestamp']) for d in csv_data if d['type'] == 'decision']
                        csv_file.seek(0)

                        if len(dynamic_decisions) == 0:
                            final_dynamic_decision = [(d['value'], d['timestamp']) for d in csv_data if d['type'] == 'final'][0]
                        else:
                            final_dynamic_decision = dynamic_decisions[-1]

                        confidence_file_path = confidence_filename.format(file_id=file_id)

                        if not Path(d + '/' + confidence_file_path).exists():
                            print(f'Missing confidence file: {confidence_file_path}!')
                            continue

                        confidence = (0, 0)

                        with open(d + '/' + confidence_file_path, 'r') as conf_csv_file:
                            conf_csv_data = csv.DictReader(conf_csv_file)
                            confidence = [(d['value'], d['timestamp']) for d in conf_csv_data if d['type'] == 'final'][0]

                        # Look for the matching dichotomous file
                        lock_in_file_path = final_decision_filename.format(file_id=file_id)

                        if not Path(d + '/' + lock_in_file_path).exists():
                            print(f'Missing lock in file: {lock_in_file_path}!')
                            continue

                        with open(d + '/' + lock_in_file_path, 'r') as di_csv_file:
                            di_csv_data = csv.DictReader(di_csv_file)
                            dichtomous_decision = [(d['value'], d['timestamp']) for d in di_csv_data if d['type'] == 'final'][0]
                            di_csv_file.seek(0)

                            data_struct = (
                                participant_id,
                                final_dynamic_decision[0] if participant_id % 2 == 1 else None,
                                int(final_dynamic_decision[1])/1000.0,
                                dichtomous_decision[0],
                                int(dichtomous_decision[1])/1000.0,
                                confidence[0],
                                int(confidence[1])/1000.0,
                            )

                        if participant_id % 2 == 0:
                            videos['videos'][video_id][lt_str].insert(0, data_struct)
                        else:
                            videos['videos'][video_id][lt_str].append(data_struct)

    write_to_single_xlsx(videos)
    write_to_xlsx(videos)
