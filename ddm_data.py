import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
import os
import re


dynamic_filename = re.compile(r'lie_truth_dynamic_([0-9]*)\.csv')

# First match is the video id #, the second match is lie/truth (1 = lie, 2 = truth)
video_id_matcher = re.compile(r'videos/([0-9]*)/alibi([0-9])_control\.webm')

def extract_video_id_and_condition(matcher, video_path):
    video_id = None
    lie_or_truth = 0
    search = matcher.search(video_path)

    if search is not None:
        video_id = int(search.groups()[0])
        lie_or_truth = int(search.groups()[1])

    return video_id, lie_or_truth

def setup_video_struct(videos: dict, video_id: int, lie_or_truth: int) -> dict:
    if 'videos' not in videos:
        videos['videos'] = {}

    if video_id not in videos['videos'].keys():
        videos['videos'][video_id] = {}

    if lie_or_truth == 1 and 'lie' not in videos['videos'][video_id]: # Lie
        videos['videos'][video_id]['lie'] = []
    elif lie_or_truth == 2 and 'truth' not in videos['videos'][video_id]: # Truth
        videos['videos'][video_id]['truth'] = []

    return videos

def write_data_to_cells(worksheet, row, data, starting_column = 1, bold = False):
    column = starting_column
    for d in data:
        worksheet.cell(row, column, d)
        if bold:
            worksheet.cell(row, column).font = Font(bold=True)
        column += 1

def write_to_xlsx(data):
    workbook = Workbook()

    for v_id in sorted(data['videos'].keys()):
        for cond in sorted(data['videos'][v_id].keys()):
            video_worksheet = workbook.create_sheet(f'Video {v_id}{"L" if cond == "lie" else "T"}')

            row = 1
            write_data_to_cells(video_worksheet, row, ('ParticipantID', 'RT', 'Choice'), bold=True)
            for d in data['videos'][v_id][cond]:
                row += 1
                write_data_to_cells(video_worksheet, row, d)

    # Write the workbook to disk
    if not Path('reports/').exists():
        Path('reports/').mkdir()

    workbook.remove(workbook['Sheet'])
    workbook.save(f'reports/DDM.xlsx')

if __name__ == '__main__':
    print('Converting data files into XLXS files...')

    subdirs = [d[0] for d in os.walk('output/')]

    videos = {}

    for d in subdirs:
        split_path = d.split('/')

        if len(split_path) == 2 and len(split_path[1]) == 0:
            continue

        participant_id = int(split_path[1])

        if participant_id % 2 != 1:
            continue
        
        print(f'Reading participant data in "{d}"')

        # Extract all of the data from the participant files
        for f in os.listdir(d):
            if dynamic_filename.match(f):
                with open(d + '/' + f, 'r') as csv_file:
                    csv_data = csv.DictReader(csv_file)

                    path = [d['value'] for d in csv_data if d['type'] == 'path'][0]
                    csv_file.seek(0)

                    video_id, lie_or_truth = extract_video_id_and_condition(video_id_matcher, path)

                    if video_id is not None:
                        videos = setup_video_struct(videos, video_id, lie_or_truth)

                        lt_str = 'lie' if lie_or_truth == 1 else 'truth'
                        videos['videos'][video_id][lt_str].append([(participant_id, int(d['timestamp'])/1000.0, 1 if int(d['value']) < 0 else 2) for d in csv_data if d['type'] == 'final'][0])

    write_to_xlsx(videos)
